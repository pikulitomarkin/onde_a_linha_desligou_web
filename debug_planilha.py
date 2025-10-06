from openpyxl import load_workbook
from pathlib import Path

# Carrega a planilha da linha lna_assis
planilha_path = Path('static/resources/KM LNA ASS.xlsx')
if planilha_path.exists():
    workbook = load_workbook(planilha_path, data_only=True)
    sheet = workbook.active
    
    # Verifica as colunas
    colunas = [str(cell.value).strip().upper() if cell.value else "" for cell in sheet[1]]
    print('Colunas disponíveis na planilha:')
    for i, col in enumerate(colunas):
        print(f'  {i}: "{col}"')
    
    print('\nPrimeiras 10 linhas da planilha:')
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=11, values_only=True)):
        print(f'{i:2d}: {row}')
    
    # Procura por torres próximas ao KM 35
    print('\n' + '='*60)
    print('PROCURANDO TORRES PRÓXIMAS AO KM 35:')
    print('='*60)
    
    # Índices das colunas
    idx_lna_ass = colunas.index("KM - LNA - ASS") if "KM - LNA - ASS" in colunas else -1
    idx_ass_lna = colunas.index("KM - ASS - LNA") if "KM - ASS - LNA" in colunas else -1
    idx_codigo = colunas.index("CODIGO") if "CODIGO" in colunas else -1
    
    print(f'Índice coluna "KM - LNA - ASS": {idx_lna_ass}')
    print(f'Índice coluna "KM - ASS - LNA": {idx_ass_lna}')
    print(f'Índice coluna "CODIGO": {idx_codigo}')
    
    if idx_ass_lna >= 0 and idx_codigo >= 0:
        print('\nTorres próximas ao KM 35 na coluna "KM - ASS - LNA":')
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[idx_ass_lna] is not None and row[idx_codigo] is not None:
                km_val = row[idx_ass_lna]
                codigo = row[idx_codigo]
                if isinstance(km_val, (int, float)) and 30 <= km_val <= 40:
                    print(f'  KM {km_val:5.1f} -> Torre: {codigo}')
    
    if idx_lna_ass >= 0 and idx_codigo >= 0:
        print('\nTorres próximas ao KM 35 na coluna "KM - LNA - ASS":')
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[idx_lna_ass] is not None and row[idx_codigo] is not None:
                km_val = row[idx_lna_ass]
                codigo = row[idx_codigo]
                if isinstance(km_val, (int, float)) and 30 <= km_val <= 40:
                    print(f'  KM {km_val:5.1f} -> Torre: {codigo}')
                    
else:
    print('Arquivo KM LNA ASS.xlsx não encontrado')
